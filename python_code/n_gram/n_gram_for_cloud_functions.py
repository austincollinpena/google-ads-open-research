import logging

import pandas as pd
import os
import numpy as np
from n_gram.n_gram_utils import basic_clean, create_dataframe_of_ngram_stats, create_dataframe_of_ngram_stats_positive_and_negative
import nltk
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tempfile import NamedTemporaryFile
from interface_with_gcp.cloud_storage.upload_file import upload_file
import time
import uuid
from nltk import word_tokenize
from nltk.util import everygrams
from pandarallel import pandarallel


def p2f(x):
    if x == " --":
        return 0
    return float(x.strip('%')) / 100


# prepare_data reads the csv document and returns with the correct columns
def prepare_data(data_for_analysis: str) -> pd.DataFrame:
    search_term_data = pd.read_csv(data_for_analysis, thousands=",", converters={'Impr. (Top) %': p2f, 'Impr. (Abs. Top) %': p2f})
    columns = ['Search term', 'Campaign', 'Ad group', 'Clicks', 'Impr.', 'Cost', 'Conversions', 'Conv. value', 'Impr. (Top) %', 'Impr. (Abs. Top) %']
    search_term_data = search_term_data[columns]
    search_term_data = search_term_data[search_term_data['Impr.'] > 6]
    search_term_data.dropna(subset=['Impr.'], inplace=True)
    return search_term_data


# Generate grams gets a list of grams by campaign
def generate_grams(search_term_data):
    campaign_names = search_term_data["Campaign"].unique()
    campaign_names = np.append(campaign_names, "all_campaigns")
    all_grams = {}
    for campaign_name in campaign_names:
        # all campaigns should make a vec of everything
        if campaign_name == "all_campaigns":
            df_of_current_campaign = search_term_data.copy(deep=True)
        else:
            df_of_current_campaign = search_term_data.loc[search_term_data["Campaign"] == campaign_name]

        words = basic_clean(''.join(str(df_of_current_campaign['Search term'].tolist())))

        gram_vecs = []
        for x in range(1, 8):
            current_grams = nltk.ngrams(words, x)
            for gram in current_grams:
                current_word = " ".join(gram)
                gram_vecs.append(current_word)

        gram_vecs = set(gram_vecs)
        all_grams[campaign_name] = gram_vecs

    return all_grams


# create_negatived_frame makes a dataframe copy with low ROAS terms removed
def create_negatived_frame(roas_target, search_term_data):
    search_term_data['roas'] = search_term_data['Conv. value'] / search_term_data['Cost']
    low_value_search_terms_excluded = search_term_data[search_term_data['roas'] > roas_target]
    return low_value_search_terms_excluded


def execute_ngrams(search_term_data, all_grams, low_value_search_terms_excluded):
    df_columns = ['gram_count', 'count', 'Impr.', 'Clicks', 'cpc', 'Cost', 'Conversions', 'Conv. value', 'top_impressions', 'abs_top_impressions']
    ngram_analysis_dataframes = {}

    def clean_dataframe(campaign_name):
        # Debug because all_campaigns isn't ready to be set
        if campaign_name == "all_campaigns":
            return
        copy = ngram_analysis_dataframes[campaign_name]
        # clean the new data frame
        ngram_analysis_dataframes[campaign_name] = ngram_analysis_dataframes[campaign_name].round(2)
        ngram_analysis_dataframes[campaign_name].index.rename('Search term', inplace=True)
        ngram_analysis_dataframes[campaign_name]['gram_count'] = ngram_analysis_dataframes[campaign_name].index.str.count(" ") + 1
        ngram_analysis_dataframes[campaign_name].replace(0, np.nan, inplace=True)
        ngram_analysis_dataframes[campaign_name].dropna(subset=['Conversions'])
        ngram_analysis_dataframes[campaign_name]['ctr'] = ngram_analysis_dataframes[campaign_name]['Clicks'] / ngram_analysis_dataframes[campaign_name]['Impr.']
        ngram_analysis_dataframes[campaign_name]['cpc'] = ngram_analysis_dataframes[campaign_name]['Cost'] / ngram_analysis_dataframes[campaign_name]['Clicks']
        ngram_analysis_dataframes[campaign_name]['cpa'] = ngram_analysis_dataframes[campaign_name]['Cost'] / ngram_analysis_dataframes[campaign_name]['Conversions']
        ngram_analysis_dataframes[campaign_name]['roas'] = ngram_analysis_dataframes[campaign_name]['Conv. value'] / ngram_analysis_dataframes[campaign_name]['Cost']

        ngram_analysis_dataframes[campaign_name]['roas_efficient'] = ngram_analysis_dataframes[campaign_name]['Conv. value_efficient'] / ngram_analysis_dataframes[campaign_name][
            'Cost_efficient']
        ngram_analysis_dataframes[campaign_name]['top_is'] = ngram_analysis_dataframes[campaign_name]['top_impressions'] / ngram_analysis_dataframes[campaign_name]['Impr.']
        ngram_analysis_dataframes[campaign_name]['abs_top_is'] = ngram_analysis_dataframes[campaign_name]['abs_top_impressions'] / ngram_analysis_dataframes[campaign_name][
            'Impr.']
        ngram_analysis_dataframes[campaign_name]["campaign_name"] = campaign_name

    # Loop through the campaign names and create the preconditions for parallel processing

    for campaign_name in all_grams:

        # Filter the DF to only the current campaign
        if campaign_name == "all_campaigns":
            campaign_df = search_term_data
            campaign_df_efficient = low_value_search_terms_excluded
        else:
            campaign_df = search_term_data.loc[search_term_data["Campaign"] == campaign_name]
            campaign_df_efficient = low_value_search_terms_excluded[low_value_search_terms_excluded["Campaign"] == campaign_name]

        # Turn the current dataframe into a dict
        all_search_term_data_dict = campaign_df.to_dict('records')
        all_search_term_data_dict_efficient = campaign_df_efficient.to_dict('records')

        # get stats from the n gram (all_grams[campaign_name] gets the n grams)
        dict_with_ngram_stats = create_dataframe_of_ngram_stats(all_search_term_data_dict, all_grams[campaign_name])
        dict_with_ngram_stats_efficient = create_dataframe_of_ngram_stats(all_search_term_data_dict_efficient, all_grams[campaign_name])

        # create the dataframes
        ngram_analysis_from_search_terms = pd.DataFrame.from_dict(dict_with_ngram_stats, orient='index', columns=df_columns)
        ngram_analysis_from_search_terms = ngram_analysis_from_search_terms.reset_index().rename(columns={'index': 'Search term'})

        ngram_analysis_from_search_terms_efficient = pd.DataFrame.from_dict(dict_with_ngram_stats_efficient, orient='index', columns=df_columns)
        ngram_analysis_from_search_terms_efficient = ngram_analysis_from_search_terms_efficient.reset_index().rename(columns={'index': 'Search term'})
        # merge the dataframes

        merged = ngram_analysis_from_search_terms.merge(ngram_analysis_from_search_terms_efficient, left_on='Search term', right_on='Search term', how='left',
                                                        suffixes=['', '_efficient'])

        merged.set_index('Search term', inplace=True)

        # Create a new dataframe
        ngram_analysis_dataframes[campaign_name] = merged

        clean_dataframe(campaign_name)

    ## Make a doc for the entire account
    account_search_term_data_dict = search_term_data.to_dict('records')
    dict_with_account_ngram_stats = create_dataframe_of_ngram_stats(account_search_term_data_dict, all_grams['all_campaigns'])
    ngram_analysis_dataframes['all_campaigns'] = pd.DataFrame.from_dict(dict_with_account_ngram_stats, orient='index', columns=df_columns)
    clean_dataframe('all_campaigns')

    all_ngram_data = pd.concat(list(ngram_analysis_dataframes.values()))
    all_ngram_data.sort_values(by='count', ascending=False, inplace=True)
    return all_ngram_data


def create_excel_file(all_ngram_data, account_name, save_locally=True):
    # Get the max gram counts so I know what to loop through
    max_gram_count = all_ngram_data["gram_count"].max()
    all_campaigns = all_ngram_data["campaign_name"].unique().tolist()

    wb = Workbook()

    for campaign in all_campaigns:
        if not isinstance(campaign, str):
            continue
        for x in range(1, max_gram_count):
            # Filter the current dataframe
            just_current_gram = all_ngram_data.loc[(all_ngram_data["gram_count"] == x)
                                                   & (all_ngram_data["campaign_name"] == campaign)].copy(deep=True)
            just_current_gram.sort_values(by=['count'], ascending=False, inplace=True)
            name = ""
            if len(campaign) > 20:
                name = f'{campaign[0:19]}-{x}-gram'
            else:
                name = f'{campaign}-{x}-gram'
            name = name.replace("/", "-")
            current_sheet = wb.create_sheet(name)
            rows = dataframe_to_rows(just_current_gram)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    current_sheet.cell(row=r_idx, column=c_idx, value=value)

    print("saving")
    if save_locally:
        wb.save(f'./n_gram/git_ignored_data/processed_files/{account_name}.xlsx')
    else:
        with NamedTemporaryFile(delete=False) as tmp:
            print("saving2")
            wb.save(tmp.name)
            tmp.seek(0)
            file_name = f'{account_name}-{uuid.uuid4()}.xlsx'
            upload_file(file_name, tmp, 'access-cloud-storage-buckets', 'temporary-ads-data-storage')


def vector_generate_exploded_grams(df: pd.DataFrame) -> pd.DataFrame:
    pandarallel.initialize()
    df['n_gram'] = df['Search term'] \
        .str.replace(",", "") \
        .str.replace(".", "") \
        .str.replace("!", "") \
        .str.replace("?", "") \
        .str.replace(":", "") \
        .parallel_apply(
        lambda x: list(map(" ".join, everygrams(word_tokenize(x), max_len=6)))
    )
    return df.explode('n_gram')


def group_campaign_level_data(df: pd.DataFrame) -> pd.DataFrame:
    campaign_grouped = group_df_on_grams(df,
                                         ['n_gram', 'Campaign'])
    return campaign_grouped


def group_df_on_grams(df: pd.DataFrame, group_on_columns: list) -> pd.DataFrame:
    return df.groupby(group_on_columns, as_index=False).agg(
        Clicks=("Clicks", "sum"),
        conversion_value=("Conv. value", "sum"),
        conversions=("Conversions", "sum"),
        cost=("Cost", "sum"),
        impressions=("Impr.", "sum"),
        n_gram_count=("n_gram", "count")
    )


def n_gram_for_cloud_functions(roas_target, data_for_analysis, account_name, save_locally):
    try:
        tic = time.perf_counter()

        search_term_data = prepare_data(data_for_analysis)
        exploded_ngram_dataframe = vector_generate_exploded_grams(search_term_data)
        campaign_grouped = group_campaign_level_data(exploded_ngram_dataframe)
        campaign_grouped.sort_values(by=['cost'], inplace=True, ascending=False)
        campaign_grouped.head(50)
        #
        # all_grams = generate_grams(search_term_data)
        # low_value_search_terms_excluded = create_negatived_frame(roas_target, search_term_data)
        # all_ngram_data = execute_ngrams(search_term_data, all_grams, low_value_search_terms_excluded)
        # create_excel_file(all_ngram_data, account_name, save_locally)
        toc = time.perf_counter()
        print(f"Finished in {toc - tic:0.4f} seconds")
    except Exception as e:
        print(e)
        logging.exception(e)


# min 6 impressions
# 19.5021

if __name__ == "__main__":
    os.chdir("../")
    n_gram_for_cloud_functions(2.2, "./n_gram/git_ignored_data/search_terms.csv", "prolock", True)
