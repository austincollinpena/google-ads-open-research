import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import TypedDict
from tempfile import NamedTemporaryFile
import uuid
from interface_with_gcp.cloud_storage.upload_file import upload_file
from pyexcelerate import Workbook
import random


class DFAndName(TypedDict):
    df: pd.DataFrame
    campaign: str
    gram_count: int


def save_multiple_dataframes_into_excel_pyexcelerate(dfs: list[pd.DataFrame], save_locally: bool, account_name: str) -> str:
    file_name = f'{account_name}-{uuid.uuid4()}.xlsx'
    fast_wb = Workbook()
    with NamedTemporaryFile(delete=False, prefix='xlsx') as tmp:

        for df in dfs:
            if df.empty:
                continue

            campaign_name = df.iloc[0]['Campaign']
            gram_count = df.iloc[0]['gram_count']
            df = df.sort_values(by=['n_gram_count'], ascending=False)
            name = ""
            if len(campaign_name) > 28:
                rand_string = ''.join(random.choice([chr(i) for i in range(ord('a'), ord('z'))]) for _ in range(3))
                name = f'{campaign_name[:24]}-{rand_string}-{gram_count}'
            else:
                name = f'{campaign_name}-{gram_count}'
            name = name.replace("/", "-")

            values = [df.columns] + list(df.values)
            fast_wb.new_sheet(name, data=values)

        if save_locally:
            fast_wb.save(file_name)
        else:
            fast_wb.save(tmp.name)

        upload_file(file_name, tmp, 'access-cloud-storage-buckets', 'temporary-ads-data-storage')
        return file_name


def save_multiple_dataframes_into_excel_pd_native(dfs: list[pd.DataFrame], save_locally: bool, account_name: str) -> str:
    file_name = f'{account_name}-{uuid.uuid4()}.xlsx'
    with NamedTemporaryFile(delete=False, prefix='xlsx') as tmp:
        if save_locally:
            writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
        else:
            writer = pd.ExcelWriter(tmp.name, engine='xlsxwriter')

        for df in dfs:
            if df.empty:
                continue

            campaign_name = df.iloc[0]['Campaign']
            gram_count = df.iloc[0]['gram_count']
            df = df.sort_values(by=['n_gram_count'], ascending=False)
            name = ""
            if len(campaign_name) > 20:
                name = f'{campaign_name[0:19]}-{gram_count}-gram'
            else:
                name = f'{campaign_name}-{gram_count}-gram'
            name = name.replace("/", "-")
            df.to_excel(writer, sheet_name=name)

        writer.save()

        upload_file(file_name, tmp, 'access-cloud-storage-buckets', 'temporary-ads-data-storage')
        return file_name


# This takes about 2x as the other implementation
def save_multiple_dataframes_into_excel_doc(dfs: list[pd.DataFrame], save_locally: bool, account_name: str) -> str:
    """
    :param dfs: list of dataframes to save
    :param save_locally: tells the function either save the workbook locally or not
    :param account_name: helps name the file
    :return:
    """
    wb = Workbook()
    for df in dfs:
        if df.empty:
            continue

        campaign_name = df.iloc[0]['Campaign']
        gram_count = df.iloc[0]['gram_count']
        df = df.sort_values(by=['n_gram_count'], ascending=False)
        name = ""
        if len(campaign_name) > 20:
            name = f'{campaign_name[0:19]}-{gram_count}-gram'
        else:
            name = f'{campaign_name}-{gram_count}-gram'
        name = name.replace("/", "-")
        current_sheet = wb.create_sheet(name)
        rows = dataframe_to_rows(df)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                current_sheet.cell(row=r_idx, column=c_idx, value=value)

    if save_locally:
        wb.save(f'./n_gram/git_ignored_data/processed_files/{account_name}.xlsx')
        return ""

    with NamedTemporaryFile(delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        tmp.name = f'{account_name}-{uuid.uuid4()}.xlsx'
        upload_file(tmp.name, tmp, 'access-cloud-storage-buckets', 'temporary-ads-data-storage')
        return tmp.name
